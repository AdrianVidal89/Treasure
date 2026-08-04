"""Vistas del Informe Hacienda: pantalla HTML + export a PDF y .xlsx."""
import datetime
from decimal import Decimal
from io import BytesIO

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.shortcuts import redirect, render

from .informe_hacienda import calcular_informe_ventas

MESES_ES = [
    '', 'ene', 'feb', 'mar', 'abr', 'may', 'jun',
    'jul', 'ago', 'sep', 'oct', 'nov', 'dic',
]

COMO_DECLARAR = [
    "Las ganancias y pérdidas por la venta de acciones, ETFs, fondos o criptomonedas "
    "se integran en la BASE IMPONIBLE DEL AHORRO del IRPF (declaración de la Renta, "
    "modelo 100).",
    "Se declaran en el apartado «Ganancias y pérdidas patrimoniales derivadas de la "
    "transmisión de elementos patrimoniales». Cada venta aporta: valor de transmisión "
    "(importe de venta − gastos) menos valor de adquisición (coste + comisiones de compra).",
    "Las pérdidas de un ejercicio se compensan con las ganancias del mismo tipo; si "
    "queda saldo negativo, puede compensarse en los 4 años siguientes.",
    "Sobre la ganancia neta se aplican los tramos de la base del ahorro (19 % / 21 % / "
    "23 % / 27 % / 28 %). La cuota mostrada es una ESTIMACIÓN orientativa.",
    "Este informe aplica el criterio FIFO (primero en entrar, primero en salir) que "
    "exige Hacienda para valores homogéneos, incluyendo las comisiones de compra en el "
    "valor de adquisición. Contrasta las cifras con tu asesor antes de presentarlas.",
    "Los intereses de DEPÓSITOS son rendimientos del capital mobiliario (no plusvalías): "
    "se declaran en su apartado propio del IRPF, pero tributan también en la base del "
    "ahorro con los mismos tramos. El interés mostrado por ejercicio es el devengado ese "
    "año; la tributación exacta depende de cuándo el banco liquida/paga el interés.",
]


def _get_hogar(request):
    profile = getattr(request.user, 'userprofile', None)
    if not profile or not profile.hogar:
        return None, None
    return profile, profile.hogar


def _anio_de(request):
    hoy = datetime.date.today()
    try:
        return int(request.GET.get('anio', hoy.year))
    except (ValueError, TypeError):
        return hoy.year


@login_required
def informe_hacienda(request):
    profile, hogar = _get_hogar(request)
    if not hogar:
        messages.error(request, "Necesitas pertenecer a un hogar.")
        return redirect('dashboard')

    anio = _anio_de(request)
    informe = calcular_informe_ventas(hogar, anio)
    hoy = datetime.date.today()
    anios = list(range(hoy.year - 5, hoy.year + 1))[::-1]

    return render(request, 'inversiones/informe_hacienda.html', {
        'informe': informe,
        'anio_actual': anio,
        'anios': anios,
        'como_declarar': COMO_DECLARAR,
    })


# ─── Export xlsx ────────────────────────────────────────────────────────────

def _fmt(valor):
    return float(valor) if isinstance(valor, Decimal) else valor


@login_required
def informe_hacienda_xlsx(request):
    profile, hogar = _get_hogar(request)
    if not hogar:
        return redirect('dashboard')

    anio = _anio_de(request)
    informe = calcular_informe_ventas(hogar, anio)

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = f"Ventas {anio}"

    encabezado_fill = PatternFill('solid', fgColor='1B4332')
    encabezado_font = Font(color='FFFFFF', bold=True)
    negrita = Font(bold=True)
    euro = '#,##0.00\\ €'

    ws['A1'] = f"Informe de plusvalías {anio} — {hogar.nombre}"
    ws['A1'].font = Font(bold=True, size=14, color='1B4332')
    ws.merge_cells('A1:K1')

    if informe.get('hay_sin_cobertura'):
        aviso = ws.cell(
            row=2, column=1,
            value=(f"Aviso: {informe['ventas_sin_cobertura']} venta(s) sin compras "
                   "suficientes registradas; su ganancia sale sobrestimada. "
                   "Ver hoja «Detalle FIFO»."),
        )
        aviso.font = Font(color='B4442E', bold=True)
        ws.merge_cells('A2:K2')

    columnas = [
        ('Fecha venta', 14), ('Activo', 26), ('Ticker', 12), ('Tipo', 12),
        ('Titular', 14), ('Cantidad', 14), ('Precio venta', 14),
        ('Importe venta', 15), ('Coste (FIFO)', 15), ('Comisión', 12),
        ('Ganancia/Pérdida', 17),
    ]
    fila_encabezado = 3
    for col, (titulo, ancho) in enumerate(columnas, start=1):
        celda = ws.cell(row=fila_encabezado, column=col, value=titulo)
        celda.fill = encabezado_fill
        celda.font = encabezado_font
        celda.alignment = Alignment(horizontal='center')
        ws.column_dimensions[get_column_letter(col)].width = ancho

    fila = fila_encabezado + 1
    for v in informe['ventas']:
        valores = [
            v['fecha'].strftime('%d/%m/%Y'), v['inversion'], v['ticker'], v['tipo_activo'],
            v['titular'], _fmt(v['cantidad']), _fmt(v['precio_venta']),
            _fmt(v['importe_venta']), _fmt(v['coste_adquisicion']), _fmt(v['comision']),
            _fmt(v['ganancia']),
        ]
        for col, valor in enumerate(valores, start=1):
            celda = ws.cell(row=fila, column=col, value=valor)
            if col in (7, 8, 9, 10, 11):
                celda.number_format = euro
        fila += 1

    # Totales
    fila += 1
    ws.cell(row=fila, column=7, value='Totales').font = negrita
    totales = {
        8: informe['total_importe'], 9: informe['total_coste'],
        10: informe['total_comisiones'], 11: informe['ganancia_neta'],
    }
    for col, valor in totales.items():
        celda = ws.cell(row=fila, column=col, value=_fmt(valor))
        celda.number_format = euro
        celda.font = negrita

    # Resumen fiscal
    fila += 2
    resumen = [
        ('Ganancias del ejercicio', informe['ganancias_positivas']),
        ('Pérdidas del ejercicio', informe['perdidas']),
        ('Ganancia neta (plusvalías)', informe['ganancia_neta']),
        ('Rendimientos de depósitos', informe['interes_depositos']),
        ('Base del ahorro', informe['base_ahorro']),
        ('Impuesto estimado (base del ahorro)', informe['impuesto_estimado']),
    ]
    for etiqueta, valor in resumen:
        ws.cell(row=fila, column=7, value=etiqueta).font = negrita
        celda = ws.cell(row=fila, column=8, value=_fmt(valor))
        celda.number_format = euro
        fila += 1

    # Hoja de detalle de adquisición FIFO: de qué compras sale el coste de cada venta.
    ws3 = wb.create_sheet('Detalle FIFO')
    cols_det = [('Fecha venta', 14), ('Activo', 24), ('Ticker', 12), ('Fecha compra', 14),
                ('Cartera / origen', 22), ('Cantidad', 14), ('Precio compra', 14),
                ('Coste (incl. comisión)', 22)]
    for c, (titulo, ancho) in enumerate(cols_det, start=1):
        celda = ws3.cell(row=1, column=c, value=titulo)
        celda.fill = encabezado_fill
        celda.font = encabezado_font
        celda.alignment = Alignment(horizontal='center')
        ws3.column_dimensions[get_column_letter(c)].width = ancho
    r = 2
    for v in informe['ventas']:
        for l in v['lotes']:
            ws3.cell(row=r, column=1, value=v['fecha'].strftime('%d/%m/%Y'))
            ws3.cell(row=r, column=2, value=v['inversion'])
            ws3.cell(row=r, column=3, value=v['ticker'])
            ws3.cell(row=r, column=4, value=l['fecha'].strftime('%d/%m/%Y'))
            ws3.cell(row=r, column=5, value=l.get('origen', ''))
            ws3.cell(row=r, column=6, value=_fmt(l['cantidad']))
            ws3.cell(row=r, column=7, value=_fmt(l['precio'])).number_format = euro
            ws3.cell(row=r, column=8, value=_fmt(l['coste'])).number_format = euro
            r += 1
        if not v['cubierto']:
            ws3.cell(row=r, column=1, value=v['fecha'].strftime('%d/%m/%Y'))
            ws3.cell(row=r, column=2, value=v['inversion'])
            c4 = ws3.cell(row=r, column=4, value='SIN COMPRA REGISTRADA')
            c4.font = Font(color='B4442E', bold=True)
            ws3.cell(row=r, column=6, value=_fmt(v['unidades_sin_coste']))
            r += 1

    # Hoja de rendimientos de depósitos
    if informe['depositos']:
        wsd = wb.create_sheet('Depósitos')
        cols_dep = [('Depósito', 26), ('Titular', 14), ('Apertura', 14), ('Tipo interés', 14),
                    ('Frecuencia', 14), (f'Interés {anio}', 16), ('Interés total', 16), ('Liquidado', 14)]
        for c, (titulo, ancho) in enumerate(cols_dep, start=1):
            celda = wsd.cell(row=1, column=c, value=titulo)
            celda.fill = encabezado_fill
            celda.font = encabezado_font
            celda.alignment = Alignment(horizontal='center')
            wsd.column_dimensions[get_column_letter(c)].width = ancho
        rd = 2
        for d in informe['depositos']:
            wsd.cell(row=rd, column=1, value=d['nombre'])
            wsd.cell(row=rd, column=2, value=d['titular'])
            wsd.cell(row=rd, column=3, value=d['apertura'].strftime('%d/%m/%Y') if d['apertura'] else '—')
            wsd.cell(row=rd, column=4, value=(float(d['tipo_interes']) / 100 if d['tipo_interes'] else 0)).number_format = '0.000%'
            wsd.cell(row=rd, column=5, value=d['frecuencia'])
            wsd.cell(row=rd, column=6, value=_fmt(d['interes_anio'])).number_format = euro
            wsd.cell(row=rd, column=7, value=_fmt(d['interes_total'])).number_format = euro
            wsd.cell(row=rd, column=8, value=d['liquidado'].strftime('%d/%m/%Y') if d['liquidado'] else '—')
            rd += 1
        wsd.cell(row=rd + 1, column=5, value='Total rendimientos').font = negrita
        celda = wsd.cell(row=rd + 1, column=6, value=_fmt(informe['interes_depositos']))
        celda.number_format = euro
        celda.font = negrita

    # Hoja de cómo declararlo
    ws2 = wb.create_sheet('Cómo declararlo')
    ws2.column_dimensions['A'].width = 110
    ws2['A1'] = f"Cómo declarar estas plusvalías ({anio})"
    ws2['A1'].font = Font(bold=True, size=13, color='1B4332')
    for i, texto in enumerate(COMO_DECLARAR, start=3):
        celda = ws2.cell(row=i, column=1, value=f"• {texto}")
        celda.alignment = Alignment(wrap_text=True, vertical='top')
        ws2.row_dimensions[i].height = 45

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    resp = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    resp['Content-Disposition'] = f'attachment; filename="informe_hacienda_{anio}.xlsx"'
    return resp


# ─── Export PDF ─────────────────────────────────────────────────────────────

@login_required
def informe_hacienda_pdf(request):
    profile, hogar = _get_hogar(request)
    if not hogar:
        return redirect('dashboard')

    anio = _anio_de(request)
    informe = calcular_informe_ventas(hogar, anio)

    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
    )

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=landscape(A4),
        topMargin=18 * mm, bottomMargin=16 * mm, leftMargin=14 * mm, rightMargin=14 * mm,
        title=f'Informe Hacienda {anio}',
    )
    estilos = getSampleStyleSheet()
    verde = colors.HexColor('#1B4332')
    estilo_titulo = ParagraphStyle('titulo', parent=estilos['Title'], textColor=verde, fontSize=18)
    estilo_h2 = ParagraphStyle('h2', parent=estilos['Heading2'], textColor=verde, fontSize=12)
    estilo_normal = ParagraphStyle('cuerpo', parent=estilos['Normal'], fontSize=9, leading=13)

    elementos = [
        Paragraph(f'Informe de plusvalías {anio}', estilo_titulo),
        Paragraph(f'{hogar.nombre} · generado el {datetime.date.today().strftime("%d/%m/%Y")}', estilo_normal),
        Spacer(1, 8 * mm),
    ]

    def moneda(v):
        return f"{float(v):,.2f} €".replace(',', 'X').replace('.', ',').replace('X', '.')

    def cant(v):
        return f"{float(v):.4f}".rstrip('0').rstrip('.')

    if informe.get('hay_sin_cobertura'):
        estilo_aviso = ParagraphStyle('aviso', parent=estilo_normal, textColor=colors.HexColor('#B4442E'))
        elementos.append(Paragraph(
            f"Aviso: {informe['ventas_sin_cobertura']} venta(s) sin compras suficientes "
            "registradas; para esas unidades no consta coste de adquisición, así que su "
            "ganancia sale sobrestimada. Registra las compras que falten en Inversiones.",
            estilo_aviso,
        ))
        elementos.append(Spacer(1, 4 * mm))

    if informe['ventas']:
        cabecera = ['Fecha', 'Activo', 'Ticker', 'Titular', 'Cant.', 'P. venta',
                    'Importe', 'Coste (FIFO)', 'Comisión', 'Ganancia/Pérdida']
        filas = [cabecera]
        for v in informe['ventas']:
            filas.append([
                v['fecha'].strftime('%d/%m/%Y'),
                v['inversion'][:22], v['ticker'], v['titular'][:14],
                f"{float(v['cantidad']):.4f}".rstrip('0').rstrip('.'),
                moneda(v['precio_venta']), moneda(v['importe_venta']),
                moneda(v['coste_adquisicion']), moneda(v['comision']),
                moneda(v['ganancia']),
            ])
        filas.append(['', '', '', '', '', 'TOTAL',
                      moneda(informe['total_importe']), moneda(informe['total_coste']),
                      moneda(informe['total_comisiones']), moneda(informe['ganancia_neta'])])

        tabla = Table(filas, repeatRows=1)
        estilo_tabla = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), verde),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('ALIGN', (4, 0), (-1, -1), 'RIGHT'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#F0F3F1')]),
            ('LINEBELOW', (0, 0), (-1, 0), 0.5, verde),
            ('LINEABOVE', (0, -1), (-1, -1), 0.7, verde),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ])
        # Colorea ganancias/pérdidas de cada fila.
        for i, v in enumerate(informe['ventas'], start=1):
            color = colors.HexColor('#2D6A4F') if v['ganancia'] >= 0 else colors.HexColor('#B4442E')
            estilo_tabla.add('TEXTCOLOR', (-1, i), (-1, i), color)
        tabla.setStyle(estilo_tabla)
        elementos.append(tabla)

        # Detalle de adquisición (FIFO): de qué compras sale el coste de cada venta.
        elementos.append(Spacer(1, 8 * mm))
        elementos.append(Paragraph('Detalle de adquisición (FIFO)', estilo_h2))
        for v in informe['ventas']:
            elementos.append(Paragraph(
                f"Venta {v['fecha'].strftime('%d/%m/%Y')} · {v['inversion']} "
                f"({v['ticker'] or '—'}) — coste de adquisición {moneda(v['coste_adquisicion'])}",
                estilo_normal,
            ))
            det = [['Fecha compra', 'Cartera / origen', 'Cantidad', 'Precio compra', 'Coste']]
            for l in v['lotes']:
                det.append([l['fecha'].strftime('%d/%m/%Y'), (l.get('origen') or '—')[:26],
                            cant(l['cantidad']), moneda(l['precio']), moneda(l['coste'])])
            if not v['cubierto']:
                det.append(['sin compra', '—', cant(v['unidades_sin_coste']),
                            'no registrada', moneda(0)])
            tabla_det = Table(det, colWidths=[26 * mm, 42 * mm, 24 * mm, 30 * mm, 32 * mm])
            tabla_det.setStyle(TableStyle([
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#F0F3F1')),
                ('ALIGN', (2, 0), (-1, -1), 'RIGHT'),
                ('LINEBELOW', (0, 0), (-1, 0), 0.3, colors.HexColor('#CDD5CF')),
                ('TOPPADDING', (0, 0), (-1, -1), 2),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
            ]))
            elementos.append(tabla_det)
            elementos.append(Spacer(1, 3 * mm))
    else:
        elementos.append(Paragraph('No hay ventas registradas en este ejercicio.', estilo_normal))

    # Rendimientos de depósitos
    if informe['depositos']:
        elementos.append(Spacer(1, 8 * mm))
        elementos.append(Paragraph('Rendimientos de depósitos', estilo_h2))
        dep_filas = [['Depósito', 'Titular', 'Apertura', 'Tipo', f'Interés {anio}', 'Interés total']]
        for d in informe['depositos']:
            dep_filas.append([
                d['nombre'][:26], d['titular'][:14],
                d['apertura'].strftime('%d/%m/%Y') if d['apertura'] else '—',
                (f"{float(d['tipo_interes']):.3f}%" if d['tipo_interes'] else '0%'),
                moneda(d['interes_anio']), moneda(d['interes_total']),
            ])
        dep_filas.append(['', '', '', '', 'TOTAL', moneda(informe['interes_depositos'])])
        tabla_dep = Table(dep_filas, repeatRows=1)
        tabla_dep.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), verde),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('ALIGN', (3, 0), (-1, -1), 'RIGHT'),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        elementos.append(tabla_dep)

    elementos.append(Spacer(1, 8 * mm))
    elementos.append(Paragraph('Resumen fiscal', estilo_h2))
    resumen_filas = [
        ['Ganancias del ejercicio', moneda(informe['ganancias_positivas'])],
        ['Pérdidas del ejercicio', moneda(informe['perdidas'])],
        ['Ganancia neta (plusvalías)', moneda(informe['ganancia_neta'])],
        ['Rendimientos de depósitos', moneda(informe['interes_depositos'])],
        ['Base del ahorro', moneda(informe['base_ahorro'])],
        ['Impuesto estimado (base del ahorro)', moneda(informe['impuesto_estimado'])],
    ]
    tabla_resumen = Table(resumen_filas, colWidths=[80 * mm, 45 * mm])
    tabla_resumen.setStyle(TableStyle([
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('LINEBELOW', (0, 0), (-1, -2), 0.25, colors.HexColor('#CDD5CF')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    elementos.append(tabla_resumen)

    elementos.append(Spacer(1, 8 * mm))
    elementos.append(Paragraph('Cómo declararlo a Hacienda', estilo_h2))
    for texto in COMO_DECLARAR:
        elementos.append(Paragraph(f'• {texto}', estilo_normal))
        elementos.append(Spacer(1, 2 * mm))

    doc.build(elementos)
    buffer.seek(0)
    resp = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    resp['Content-Disposition'] = f'attachment; filename="informe_hacienda_{anio}.pdf"'
    return resp
